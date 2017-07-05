import os

import os.path
import scipy.stats

import math
import numpy as np
name = 'data.xlsx'

import openpyxl
import copy

basePath = 'PLOT/'

from scipy.stats import wilcoxon, ttest_ind,mannwhitneyu as mwit2,ttest_rel
from scipy.stats import chisqprob,norm

from openpyxl import Workbook

from openpyxl.styles import *
from openpyxl import *

import matplotlib.pyplot as plt

import matplotlib.patches as mpatches
import matplotlib
from matplotlib.patches import Rectangle,Polygon


sigLevel = 0.05

def is_num(val):
    try:
        v = float(val)
        return True
    except Exception as e:
        return False



wb = openpyxl.load_workbook(name)
ws = wb.active
minrow = 3
maxrow  = 2822

colmeasu = 5
colwc = 8
 
m = 14

mcolnof20 = m+18
mcolcost = m+6


cols = {
'auc' : m,
'f-measure' :m+1,
'recall' : m+3,
'precision' : m+2,
'pf' : m+4,
'auc-ce' : m+5,
'balance' : m+7,
'accuracy' : m+8,
'g-mean1' : m+9,
'g-mean2' : m+10,
'g-measure' : m+11,
'specificity' : m+12,
'mcc' : m+13,
'h-measure' : m+14,
'g-mean3' : m+15,
'hubertstat':m+17
}



papers = {}
preval = ''

def getpaper(row, papers):
    for t in papers.keys():
        if row>=papers[t]['startrow'] and row<=papers[t]['lastrow']:
            break
    return t

def getPaperCiteId(s,papers):
    if not s in papers.keys():
        return s
    id = int(s[1:])
    return '[S' +str(id)+']'
    

def getColVal (row,col,papers):
    val = ws.cell(row=row,column=col).value
    
    if val == None:
        p = getpaper(row,papers)
        r = row-1
        strow = papers[p]['startrow']
        while r>=strow:
            val = ws.cell(row=r,column=col).value
            if val!=None:
                return val
            r-=1

    return val


#Detect the range for each paper
for row in range(minrow,maxrow):
    val = ws.cell(row=row,column=1).value
    if val==None:
        continue

    papers[val] = {}
    papers[val]['startrow'] = row
    if preval!='':
        papers[preval]['lastrow'] = row-1
    preval = val

papers[preval]['lastrow'] = maxrow

for p in papers.keys():
    for i in range(papers[p]['startrow'],papers[p]['lastrow']+1):
        papers[p]['row'+str(i)] = {}

lstpapers = sorted(list(papers.keys()))
print (lstpapers)

measuresSet = set()

tmpSet = set()

invRows = set()

evcriteria = {}
for c in cols.keys():
    evcriteria[c] = [set(),set()]


#Detect the evaluation measures for each paper
for row in range(minrow,maxrow):
    paper = getpaper(row, papers)
    ds = ws.cell(row=row,column=7).value
    lrnr = ws.cell(row=row,column=6).value
    wc = ws.cell(row=row,column=8).value
    if ds == None and lrnr == None and wc == None:
        invRows.add(row)
    val = ws.cell(row=row,column=5).value
    if val==None:
        continue
    valc = val.replace(' ','').replace('\n','').lower().replace('*','').split(',')
    val = val.replace(' ','').replace('\n','').lower().split(',')
    
    for v in valc:
        if v.startswith('*'): 
            continue
        if not v in measuresSet:
            measuresSet.add(v)

    
    papers[paper]['measures'] = valc
    papers[paper]['measuresFull'] = val


colslst = sorted(list(cols.keys()))
dsdic = {}
allvalTypes = set()

#Extract the values from the Excel file
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    
    paper = getpaper(row, papers)
    valType = ws.cell(row=row,column=11).value
    ds = ws.cell(row=row,column=7).value
    WC = ws.cell(row=row,column=8).value
    lrnr = ws.cell(row=row,column=6).value
    dpVarLevel = getColVal(row,2,papers)
    dsSuite = getColVal(row,13,papers)
    

    if ds == None:
        
        ds = getColVal(row,7,papers)
        
    if ds !=None:
        ds = str(ds).lower()
    else:
        print ('ds None.',row)
        input('')
    if ds.startswith('median values are given'):
        if ds.find(':')>=0:
            ds = 'mdn-row:'+ds[ds.find(':')+1:]+'--'+str(row)
        else:
            ds = 'mdn-row--'+str(row)
    if ds.startswith('mean values are given'):
        if ds.find(':')>=0:
            ds = 'mn-row:'+ds[ds.find(':')+1:]+'--'+str(row)
        else:
            ds = 'mn-row'+str(row)

    
    if ds.lower().find('median')>=0 or ds.lower().find('mean')>=0:
        print (ds)
        print (row)
        input('Error')


    ds1 = ds
    if ds.find('-->')>=0:
        ds1 = ds[ds.find('-->')+3:].strip()

    if len(ds1)>0:
        papers[paper]['row'+str(row)]['ds']=ds1

    if dsSuite!=None:
        papers[paper]['row'+str(row)]['dsSuite']=dsSuite
    else:
        print (row, 'DS Suite is None')
        input('')
    metrics = ws.cell(row=row,column=4).value
    pprappr = ws.cell(row=row,column=9).value  #Data, Learner, Metric, None
    dataappr = ws.cell(row=row,column=10).value  #data approach: Filtering, Transformation, Mixed Data, Feature Selection, ...
    
    if dpVarLevel!=None:
        papers[paper]['row'+str(row)]['dpVarLevel']=dpVarLevel.lower()
        
    else:
        print('DP var level is None', row)
        input('')

    if valType!=None:
        valType = valType.lower()
        papers[paper]['row'+str(row)]['valueType']=valType
        if not valType in allvalTypes:
            allvalTypes.add(valType)
    else:
        print('Val Type is None',row)
        input('')

    if WC!=None:
        papers[paper]['row'+str(row)]['WC']=WC
    else:
        print('WC is None',row)
        input('')


    if metrics!=None:
        metrics = metrics.lower().replace('+ ','+').replace(' +','+')
        papers[paper]['row'+str(row)]['metricsLst']=metrics.split('+')
        if metrics.find('+')>0:
            papers[paper]['row'+str(row)]['metrics']=metrics
        else:
            papers[paper]['row'+str(row)]['metrics']='+'.join(sorted(metrics.split('+')))
    else:
        print('Metrics None',row)
        input('')

    if pprappr!=None:
        pprappr=pprappr.lower()
        papers[paper]['row'+str(row)]['pprappr']=pprappr.lower()
        
    else:
        print('PPRApproach None',row)
        input('')
    

    if dataappr!=None:
        if dataappr.find('+')>0:
            dataappr = '+'.join(sorted(dataappr.lower().split('+')))
        else:
            dataappr = dataappr.lower()
        papers[paper]['row'+str(row)]['dataappr']=dataappr
        tmpSet.add(dataappr)
    else:
        papers[paper]['row'+str(row)]['dataappr']='None'
        tmpSet.add('none')


    if lrnr!=None:
        lrnr = lrnr.lower().replace('\t',' ').strip()
        papers[paper]['row'+str(row)]['lrnr']=lrnr
    else:
        print('lrnr None',row)
        input('')

    papers[paper]['row'+str(row)]['colVals'] = {}
    


    for ck in colslst:
        col = cols[ck]
        if not (ck in papers[paper]['measures']):
            continue
        val = ws.cell(row=row,column = col).value
        if val!=None:
            val = str(val).strip()
            if len(val)<=0:
                continue
            if not is_num(val):
                print (row)
                try:
                    print (val)
                except Exception as e:
                    print ('Not printable')
                input('')
                continue
            
                
            v2 = float(val)
            
            papers[paper]['row'+str(row)]['colVals'][ck] = v2
            evcriteria[ck][0].add(paper)


#Calculate additional values

for row in range(minrow,maxrow):
    if row in invRows:
        continue

    paper = getpaper(row, papers)
    if len(papers[paper]['row'+str(row)]['colVals'].keys())<=0:
        invRows.add(row)
        continue
    if 'recall' in papers[paper]['row'+str(row)]['colVals'].keys():

        rec = papers[paper]['row'+str(row)]['colVals']['recall']

        if 'precision' in papers[paper]['row'+str(row)]['colVals'].keys():
            prec = papers[paper]['row'+str(row)]['colVals']['precision']
            if not 'f-measure' in papers[paper]['row'+str(row)]['colVals'].keys():
                if prec+rec!=0:
                    papers[paper]['row'+str(row)]['colVals']['f-measure'] = (2*rec*prec)/(rec+prec)
                else:
                    papers[paper]['row'+str(row)]['colVals']['f-measure'] = 0
            if not 'g-mean1' in papers[paper]['row'+str(row)]['colVals'].keys():
                papers[paper]['row'+str(row)]['colVals']['g-mean1'] = math.sqrt(rec*prec)


        if 'f-measure' in papers[paper]['row'+str(row)]['colVals'].keys():
            if not 'precision' in papers[paper]['row'+str(row)]['colVals'].keys():
                fm = papers[paper]['row'+str(row)]['colVals']['f-measure']
                if (2*rec-fm)!=0:
                    papers[paper]['row'+str(row)]['colVals']['precision'] = fm*rec/(2*rec-fm)
                else:
                    papers[paper]['row'+str(row)]['colVals']['precision'] = 0

        if 'precision' in papers[paper]['row'+str(row)]['colVals'].keys():
            prec = papers[paper]['row'+str(row)]['colVals']['precision']
            if not 'g-mean1' in papers[paper]['row'+str(row)]['colVals'].keys():
                papers[paper]['row'+str(row)]['colVals']['g-mean1'] = math.sqrt(rec*prec)



        if 'pf' in papers[paper]['row'+str(row)]['colVals'].keys():
            pf = papers[paper]['row'+str(row)]['colVals']['pf']
            if not 'balance' in papers[paper]['row'+str(row)]['colVals'].keys():
                papers[paper]['row'+str(row)]['colVals']['balance'] = 1-math.sqrt(pf*pf + (1-rec)*(1-rec))/math.sqrt(2)            
            if not 'g-measure' in papers[paper]['row'+str(row)]['colVals'].keys():
                if (rec+(1-pf))!=0:
                    papers[paper]['row'+str(row)]['colVals']['g-measure'] = (2*rec*(1-pf))/(rec+(1-pf))
                else:
                    papers[paper]['row'+str(row)]['colVals']['g-measure'] = 0
            if not 'g-mean2' in papers[paper]['row'+str(row)]['colVals'].keys():
                papers[paper]['row'+str(row)]['colVals']['g-mean2'] = math.sqrt(rec*(1-pf))


        if 'balance' in papers[paper]['row'+str(row)]['colVals'].keys():
            if not 'pf' in papers[paper]['row'+str(row)]['colVals'].keys():
                bal = papers[paper]['row'+str(row)]['colVals']['balance']                
                papers[paper]['row'+str(row)]['colVals']['pf'] = math.sqrt((2*(1-bal)*(1-bal))-((1-rec)*(1-rec)))
                
                        
        if 'pf' in papers[paper]['row'+str(row)]['colVals'].keys():
            pf = papers[paper]['row'+str(row)]['colVals']['pf']
            if not 'g-measure' in papers[paper]['row'+str(row)]['colVals'].keys():
                if (rec+(1-pf))!=0:
                    papers[paper]['row'+str(row)]['colVals']['g-measure'] = (2*rec*(1-pf))/(rec+(1-pf))
                else:
                    papers[paper]['row'+str(row)]['colVals']['g-measure'] = 0
            if not 'g-mean2' in papers[paper]['row'+str(row)]['colVals'].keys():
                papers[paper]['row'+str(row)]['colVals']['g-mean2'] = math.sqrt(rec*(1-pf))
        
                    
                    
    if 'precision' in papers[paper]['row'+str(row)]['colVals'].keys():
        prec = papers[paper]['row'+str(row)]['colVals']['precision']
        if 'f-measure' in papers[paper]['row'+str(row)]['colVals'].keys():
            if not 'recall' in papers[paper]['row'+str(row)]['colVals'].keys():
                fm = papers[paper]['row'+str(row)]['colVals']['f-measure']
                if (2*prec-fm)!=0:
                    papers[paper]['row'+str(row)]['colVals']['recall'] = fm*prec/(2*prec-fm)
                else:
                    papers[paper]['row'+str(row)]['colVals']['recall'] = 0


tset = set()

dicrows = {}

#Dict of the data rows, rowNumber as key...
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    p = getpaper(row,papers)
    dicrows[str(row)] = papers[p]['row'+str(row)]
    dicrows[str(row)]['paper'] = p

#Evaluation criteria after calculating additional values
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    sr = str(row)
    p = dicrows[sr]['paper']
    for c in dicrows[sr]['colVals']:
        evcriteria[c][1].add(p)


#Extract the list of the learners used in CPDP
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    p = getpaper(row,papers)
        
    t = papers[p]['row'+str(row)]['lrnr'].strip().replace('\t',' ')
    tset.add(t)

alllrnrs = sorted(list(tset))



#List of all evaluation metrics
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    
        
    t = dicrows[str(row)]['metrics'].strip()
    tset.add(t)
    
allmetrics = sorted(list(tset))
print(allmetrics)



#Extract all paper approaches
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    t = dicrows[str(row)]['pprappr'].strip()
    tset.add(t)
    
allpprappr = sorted(list(tset))
print(allpprappr)





#Extract the level of granularity  for the dependent variable
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    t = dicrows[str(row)]['dpVarLevel'].strip()
    tset.add(t)
    
alldpvarlevel = sorted(list(tset))
print(alldpvarlevel)





#Extract the data approaches
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    
    t = dicrows[str(row)]['dataappr'].strip()
    tset.add(t)
    
alldataappr = sorted(list(tset))
print(alldataappr)




#Extract all distinct data approaches
tset = set()
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    
    t = dicrows[str(row)]['dataappr'].strip().lower()
    for tt in t.split('+'):
        tset.add(tt)
 
alldataapprdist = sorted(list(tset))
print(alldataapprdist)

#Extract all DSSuites and Datasets
dsSuites = {}
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    
    dss = dicrows[str(row)]['dsSuite'].strip().lower()
    ds = dicrows[str(row)]['ds'].strip().lower()
    if not dss in dsSuites.keys():
        dsSuites[dss] = set([ds])
    else:
        dsSuites[dss].add(ds)



suitesOrder = ['softlab','nasa','jur','aeeem','relink']




#Abbreviations for the data approaches
abbrdic = {}
abbrdic['data transformation'] = 'DT'
abbrdic['filtering'] = 'F'
abbrdic['log-filter'] = 'LF'
abbrdic['re-weighting'] = 'RW'
abbrdic['re-sampling'] = 'RS'
abbrdic['mixed data'] = 'MD'
abbrdic['normalization'] = 'N'
abbrdic['none'] = 'None'
abbrdic['discretization']='D'
abbrdic['privacy']='P'
abbrdic['clustering']='C'
abbrdic['feature selection'] = 'FS'



#Dimensions for a string
def get_ax_size(fig,ax):
    bbox = ax.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
    width, height = bbox.width, bbox.height
    width *= fig.dpi
    height *= fig.dpi
    return width, height

def getWH(fig,ax,txt,rot = 90):
    
    r = fig.canvas.get_renderer()
    t = ax.text(-100, -100, txt,rotation=rot)

    bb = t.get_window_extent(renderer=r)
    width = bb.width
    height = bb.height
    return width, height

defFontSize = matplotlib.rcParams['font.size']
print ('Default Font size is:',defFontSize)

#import dunn
#from Orange import evaluation
#from Orange.evaluation import compute_CD

#Draw the violin plots
def draw(cklrnrdata,ck,name='lrnr',vType = 'Average',ml = []):

    ncols = 1
    nrows = 1
    lrns = sorted(list(cklrnrdata[ck].keys()))
    if len(lrns)<2:
        return
    
    ms = ','.join(ml)
    if len(ms)==0 or ms==',':
        ms='all'

    h = 3.4
    if len(lrns)>10:
        h=4.5
    
    if name.find('--')>0 and len(lrns)>=18:        
        nrows = 2
        h = h*2
        cnt = int(len(lrns)/2)
        splc = [cnt,len(lrns)-cnt]
        w = max(5,2*len(lrns)/2)
        ac = 1
    else:
        splc = [len(lrns)]
        w = max(5,2*len(lrns))
        ac = 0

    fig,axes = plt.subplots(ncols = ncols,nrows = nrows,figsize=(w,h))
    if nrows==1:
        axes = [axes]
        
    
    for ax in axes:
        
        astart = 0
        aend = 0
        if ac>0:
            astart = splc[ac-1]
            aend = splc[ac-1]+splc[ac]
        else:
            aend = splc[ac]


        ac-=1

        if aend-astart>10:
            matplotlib.rcParams.update({'font.size': 18})
        else:
            matplotlib.rcParams.update({'font.size': defFontSize})
    
        data = []
    
        btm = max([len(lrnr) for lrnr in lrns])

        for lrnr in lrns:
            data.append(cklrnrdata[ck][lrnr])
    
        s = [np.median(data[i]) for i in range(len(data))]
        inds = sorted(range(len(s)), key=lambda k: s[k])
        data2 = [data[ind] for ind in inds]
        lbls = [lrns[ind]+' ('+str(len(data[ind]))+')' for ind in inds]

        data2 = data2[astart:aend]
        lbls = lbls[astart:aend]
        mn = np.min([np.min(data2[i]) for i in range(len(data2))])
        mn = np.floor(min(0,mn)*10)/10
        ax.set_ylim(mn,1.0)
        ax.set_yticks(np.arange(mn,1.,0.1))
    
        #ax.set_xticks([])
        ax.tick_params(
            axis='x',which='both',bottom='off',top='off',labelbottom='off') 



        ylbl = vType+' '+ck 
        if nrows>1:
            ylbl+= ' ('+str(ac+1)+')'
        ax.set_ylabel(ylbl)
    
        ax.set_xlim(0,len(data2)+1)
        pos = 1
        #Place the labels (some might be long. find a correct place)
        for lbl in lbls:
            wt,ht = getWH(fig,ax,lbl)
            w,h = get_ax_size(fig,ax)
            sz = ht/h
            posy = np.mean(data2[pos-1])
            posymin = posy - sz/2
            posymax = posy + sz/2

            if posymin<0:            
                posymax-=posymin
                posymax+=0.05
                posymin=0.05
            if posymax>0.95:
                posymin -= posymax-0.95
                posymax=0.95
            if posymin<=0:
                posymax-= posymin
                posymax+=0.02
                posymin = 0.02
        
            ax.text(pos-0.4, posymax, lbl, rotation=90)
            pos+=1

        prt=ax.violinplot(data2,showmeans=True,showmedians=True)

        #Tests: Kruskal-Wallis H or Mann-Wittney U for multiple and 2 samples.
        #Not including the test results at the moment.
        if len(data2)>2:
            test = 'KW-H'
            ps = scipy.stats.mstats.kruskalwallis(*data2) 
            pv = round(ps[1],3)
            #ax.text(-0.2, 1.01, ' %s: p-value%s, statistic=%.3f' % (test,'='+str(pv) if pv>0 else '<<0.0001' ,round(ps[0],3)))
        else:
            if len(data2) == 2:
                ps = scipy.stats.mannwhitneyu(data2[0],data2[1])
                #ax.text(-0.2, 1.01, ' ManW U: p-value=%f, statistic=%f' % (round(ps[1],3),round(ps[0],3)))

        #Mean and Median markers
        prt['cmeans'].set_facecolor('black')
        prt['cmeans'].set_edgecolor('black')
        prt['cmeans'].set_linestyle('--')
        prt['cmeans'].set_linewidth(3)

    #Create Necessary dirs and Save the results
    if not os.path.exists(basePath+name+'/'+ms+'/'+vType+'/'):
        os.makedirs(basePath+name+'/'+ms+'/'+vType+'/')
    if not os.path.exists(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/'):
        os.makedirs(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/')
    if not os.path.exists(basePath+'allin1/'):
        os.makedirs(basePath+'allin1/')
    if not os.path.exists(basePath+'jpg/allin1/'):
        os.makedirs(basePath+'jpg/allin1/')
    fig.tight_layout()
    fig.savefig(basePath+name+'/'+ms+'/'+vType+'/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.eps', format='eps', dpi=1000)
    fig.savefig(basePath+'allin1/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.eps', format='eps', dpi=1000)
    
    fig.savefig(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/'+ms+'-'+'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.jpg')
    fig.savefig(basePath+'jpg/allin1/'+ms+'-'+'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.jpg')
    plt.close('all')
    
    print (ck,name, vType, 'done')

import matplotlib.gridspec as gridspec

#Draw the forest plots
def drawForstPlot(data,ck,name='lrnr',vType = 'Average',ml = [],maxX = 7.0,Q = 0,qpval = 0,taw2=0,fws = None,sfixed = None):
    showQ = False
    if Q<0 or qpval<0:
        showQ = False
    else:
        showQ=True
    
    ms = ','.join(ml)
    if len(ms)==0 or ms==',':
        ms='all'
    
    xlen = 8.0
    mxrow = -10000
    mnrow = 10000
    for i in range(2,len(data)):
        if data[i][3]>mxrow:
            mxrow = data[i][3]
        if data[i][2]<mnrow:
            mnrow = data[i][2]

    xsz = 1
    xlen = 2*maxX+xsz     
    left = 0.7
    ystart = 0
    step = 0.8
    ylen = np.max([len(data)*step])+2
    xd2 = (xlen-xsz)/2 
    xmargin = 0 
    ymargin = 0
    
    if maxX>4:
        matplotlib.rcParams.update({'font.size': 18})
    else:
        matplotlib.rcParams.update({'font.size': defFontSize})
    
    #Create the figure
    plt.figure(figsize=(xlen,ylen))
    plt.subplots_adjust(wspace=0, hspace=0)
    plt.xlim(-xd2-xsz,xd2)
    plt.ylim(-2,ylen-2)

    plt.tight_layout()
    plt.plot([-(xd2-xmargin)+0.02, (xd2-xmargin)-0.02], [ystart-1, ystart-1], color='k', linestyle='-', linewidth=1)
    plt.plot([0, 0], [ystart-1, ylen - ystart-0.02], color='k', linestyle='-', linewidth=0.7)
    
    i = 9
    j=0
    while j<xd2-1:
        plt.annotate(str(round((i+1)*0.1,2)), xy=((i+1)*0.1-0.15,ystart-1-0.3))
        plt.annotate(str(round(-(i+1)*0.1,2)), xy=(-(i+1)*0.1-0.22,ystart-1-0.3))
        plt.plot([-(i+1)*0.1, -(i+1)*0.1], [ystart-1-0.05, ystart-1+0.05], color='k', linestyle='-', linewidth=1)
        plt.plot([(i+1)*0.1, (i+1)*0.1], [ystart-1-0.05, ystart-1+0.05], color='k', linestyle='-', linewidth=1)
        i+=10
        j+=1

    
    currentAxis = plt.gca()
    currentAxis.yaxis.set_visible(False)
    currentAxis.xaxis.set_visible(False)
    
    if maxX>4:
        plt.text(0.42*xlen/2, -1.65,'Favors cross project')
        plt.text(-0.7*xlen/2, -1.65,'Favors within project')
    else:
        plt.text(0.42*xlen/2, -1.60,'Favors cross project')
        plt.text(-0.79*xlen/2, -1.60,'Favors within project')
    xstr = 'Forest Plot for '+vType+' '+ck
    
    if maxX>4:
        plt.text(-1.6, -1.8, xstr)
    else:
        plt.text(-1.3, -1.8, xstr)
            
    rat = 1
    lastrowIndex = len(data)-1
    for i in range(2,lastrowIndex):
        row = data[i]
        #calculate min and max
        r1 = row[2]
        r2 = row[3]
    
        #calculate the size of the box
        size = np.sqrt(row[4])*3/30
        #print (size)
        
           
        #the length of the line
        linelen  = r2 - r1
        #Write the study name
        plt.annotate(getPaperCiteId(row[0],papers)+' ('+str(round(row[4]*100.0/data[-1][4],2))+'%)', xy=(-xd2-xmargin/2-0.8, (i)*step-0.02))
        #draw the line
        plt.plot([r1, r2], [(i)*step, (i)*step], color='k', linestyle='-', linewidth=0.5)
        #Draw the rectangle for each study.
        #currentAxis.add_patch(Rectangle((r1+linelen/2-size*rat/2  , (i)*step-size/2 ), size*rat, size,alpha=1,color='k'))
        currentAxis.add_patch(Rectangle((row[1]-size*rat/2  , (i)*step-size/2 ), size*rat, size,alpha=1,color='k',fill=True,linewidth=0.4))

        if fws!=None:
            sizef = np.sqrt(fws[i-2])*3/30
            linelen  = r2 - r1
            currentAxis.add_patch(Rectangle((row[1]-sizef*rat/2  , (i)*step-sizef/2 ), sizef*rat, sizef,alpha=1,color='k',linestyle='-',fill=False,linewidth=0.7))

    size = 0.1
    row = data[-1]
    pvall = row[-1]
    pvalltext = ''
    if pvall<0.001:
        pvalltext = 'p-val<0.001'
    else:
        pvalltext = 'p-val=%.3f'%pvall
    plt.annotate('Random Effect (%s)'%pvalltext, xy=(-xd2-xmargin/2-0.8, size/2+0.50))
    plt.annotate('95%% CI=[%.2f , %.2f]'%(row[2],row[3]), xy=(-xd2-xmargin/2-0.8, size/2+0.1))
    if sfixed:
        points = [[row[2],size/2+0.35], [row[1],size+0.1+0.35], [row[3],size/2+0.35], [row[1],-size+0.35]]
        if row[1]<0:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.40, size/2-0.05))
        else:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.35, size/2-0.05))
        plt.plot([row[1], row[1]], [-size/2+0.35,ylen-ystart], color='k', linestyle='-.', linewidth=1)
    else:
        points = [[row[2],size/2], [row[1],size+0.1], [row[3],size/2], [row[1],-size]]
        if row[1]<0:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.40, size/2-0.5))
        else:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.35, size/2-0.5))
        plt.plot([row[1], row[1]], [-size/2,ylen-ystart], color='k', linestyle='-.', linewidth=1)
    p = Polygon(points,closed=True,color='k')
    
    if showQ:
        if qpval<0.001:

            plt.text(-xd2-xmargin/2-0.8, size/2-0.30,'($\\tau^2$=%.2f, Q=%.2f, p-val<0.001)' %(taw2,Q))
        else:
            plt.text(-xd2-xmargin/2-0.8, size/2-0.30,'($\\tau^2$=%.2f, Q=%.2f, p-val=%.3f)'  %(taw2,Q,qpval))
    
    
    currentAxis.add_patch(p)
    
    
    


    if sfixed:
        row = sfixed
        pvall = row[-1]
        pvalltext = ''
        if pvall<0.001:
            pvalltext = 'p-val<0.001'
        else:
            pvalltext = 'p-val=%.3f'%pvall

        if pvall<0.001:
            pvalltext = 'p-val<0.001'
        else:
            pvalltext = 'p-val=%.3f'%pvall
        plt.annotate('Fixed Effect (%s)'%pvalltext, xy=(3, size/2-0.40))
        plt.annotate('95%% CI=[%.2f , %.2f]'%(row[2],row[3]), xy=(3, size/2-0.75))

        points = [[row[2],size/2-0.40], [row[1],size+0.1-0.40], [row[3],size/2-0.40], [row[1],-size-0.40]]
        p = Polygon(points,closed=True,color='k',fill=False,linewidth=0.5)
        #plt.annotate('Summary (%s)'%pvalltext, xy=(-xd2-xmargin/2-0.8, size/2+0.1))
        #plt.annotate('95%% Conf. Int. = [%.3f , %.3f]'%(row[2],row[3]), xy=(-xd2-xmargin/2-0.8, size/2-0.25))
        
        currentAxis.add_patch(p)
        plt.plot([row[1], row[1]], [-size/2-0.40,ylen-ystart], color='r', linestyle=':', linewidth=0.5)
        if row[1]<0:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.40, size/2-0.85))
        else:
            plt.annotate(str(round(row[1],3)), xy=(row[1]-0.35, size/2-0.85))
    
    #if row[2]<0:
    #    plt.annotate('%.3f'%row[2], xy=(row[2]-0.40, size/2))
    #else:
    #    plt.annotate('%.3f'%row[2], xy=(row[2]-0.35, size/2))
    #if row[3]<0:
    #    plt.annotate('%.3f'%row[3], xy=(row[3]+0.1, size/2))
    #else:
    #    plt.annotate('%.3f'%row[3], xy=(row[3]+0.05, size/2))
    #Necessary dirs and save the plots
    if not os.path.exists(basePath+name+'/'+ms+'/'+vType+'/'):
        os.makedirs(basePath+name+'/'+ms+'/'+vType+'/')
    if not os.path.exists(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/'):
        os.makedirs(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/')

    if not os.path.exists(basePath+'allin1Forest/'):
        os.makedirs(basePath+'allin1Forest/')
    if not os.path.exists(basePath+'jpg/allin1Forest/'):
        os.makedirs(basePath+'jpg/allin1Forest/')

    plt.savefig(basePath+name+'/'+ms+'/'+vType+'/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.eps', format='eps', dpi=1000,transparent="True", pad_inches=0,bbox_inches='tight')

    plt.savefig(basePath+'jpg/'+name+'/'+ms+'/'+vType+'/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.jpg',transparent="True", pad_inches=0,bbox_inches='tight')
    
    plt.savefig(basePath+'allin1Forest/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.eps', format='eps', dpi=1000,transparent="True", pad_inches=0,bbox_inches='tight')
    
    plt.savefig(basePath+'jpg/allin1Forest/'+ms+'-'+ 'ck'+name.replace(' ','-')+'-'+vType+'-'+ck+'.jpg',transparent="True", pad_inches=0,bbox_inches='tight')
    
    
    plt.close('all')

#Hedges correction factor
def J(m):
    return 1-3/(4*m-1)

#A term for the variance estimate
def nhat(n1,n2):
    return (n1+n2)/(n1*n2)
    
#find and item in a list
def findinlist(lst,txt):
    for i in range(len(lst)):
        if lst[i]==txt:
            return i
    return -1


#for the measures and/or their combinations
for msList in []:#[[],['precision','recall','f-measure'],['pf','recall','balance']]:#,['f-measure'],,,,['auc']   
    #Types of reporting : average, median, original
    for vType in allvalTypes:
        cklrnrdata = {}
        ckdata = {}

        #calculate for every item in msList: all or combiations
        for ck in (msList if len(msList)>0 else colslst):
            cklrnrdata[ck] = {}
            #Violin plots for learning techniques
            for lrnr in alllrnrs:
                dd = []
                for row in range(minrow,maxrow):
                     
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)): 
                        continue
                    sr = str(row)
                    if dicrows[sr]['lrnr'] == lrnr and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    cklrnrdata[ck][lrnr] = dd
            if len(cklrnrdata[ck].keys())>0:
                draw(cklrnrdata,ck,'Learning Techniques',vType,msList)
        
            
            ckdatacpy = copy.deepcopy(cklrnrdata)
        

            #Only base learners, combine the emsembles and base learners
            lrns = sorted(list(ckdatacpy[ck].keys()));
            for lrnr in lrns:
            
                if lrnr.find(' ')>0:
                    p2 = lrnr[lrnr.find(' ')+1:]
                    if p2 in ckdatacpy[ck].keys():
                        ckdatacpy[ck][p2]+=ckdatacpy[ck][lrnr]
                        del ckdatacpy[ck][lrnr]
                    else:
                        ckdatacpy[ck][p2]=ckdatacpy[ck][lrnr][:]
                        del ckdatacpy[ck][lrnr]
                
            if len(ckdatacpy[ck].keys())>0:
                draw(ckdatacpy,ck,'Base Learners',vType,msList)
        
        
                
            #Compare the Ensembles    
            ckdatacpy = copy.deepcopy(cklrnrdata)
            newdatadic = {}
            newdatadic[ck] = {}
            lrnsEns = sorted([t for t in ckdatacpy[ck].keys() if t.find(' ')>0])
            if len(lrnsEns)>0:
                baseEnslrns = set()
                for l in lrnsEns:
                    baseEnslrns.add(l[:l.find(' ')])
                baseEnslrns = sorted(list(baseEnslrns))
                for l in baseEnslrns:
                    newdatadic[ck][l]=[]
                for l in lrnsEns:

                    newdatadic[ck][l[:l.find(' ')]]+=ckdatacpy[ck][l]
            
            if len(newdatadic[ck].keys())>0:
                draw(newdatadic,ck,'Ensemble Techniques',vType,msList)
        


            #metrics and feature sets or their combinations: OO+SCM+LOC, SCM+LOC, Source Code Text, ...
            ckdata[ck] = {}    
            for mtr in allmetrics:
                dd = []           
                for row in range(minrow,maxrow):
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)): 
                        continue

                    sr = str(row)
                    if dicrows[sr]['metrics'] == mtr and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    ckdata[ck][mtr] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Metrics and Features',vType,msList)




            #Compare the paper approaches
            ckdata[ck] = {}
            for appr in allpprappr:
                dd = []            
                for row in range(minrow,maxrow):
        
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)): 
                        continue

                    sr = str(row)
                    if dicrows[sr]['pprappr'] == appr and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    ckdata[ck][appr] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Main Approach of Studies',vType,msList)        



            #Compare data approaches
            ckdata[ck] = {}
            for appr in alldataappr:
                dd = []
            
                for row in range(minrow,maxrow):
        
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):               
                        continue
                    sr = str(row)
                    if dicrows[sr]['dataappr'] == appr and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    if appr.find('+')>0:
                        appr= '+'.join([abbrdic[part.lower()] for part in appr.lower().split('+')])
                    else:

                        if appr.lower() in abbrdic.keys():
                            appr=abbrdic[appr.lower()]
                        
                    ckdata[ck][appr] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Data Approach in Studies',vType,msList)


            
            #Compare the individual data approaches   
            ckdata[ck] = {}
            for appr in alldataapprdist:
                dd = []
            
                for row in range(minrow,maxrow):
        
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):
                        continue
                    sr = str(row)
                    if findinlist(dicrows[sr]['dataappr'].lower().split('+'),appr)>=0 and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    ckdata[ck][appr] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Aggregated Data Approach in Studies',vType,msList)



            #Compare Learner-individual-data approach combinations
            ckdata[ck] = {}
            for lrnr in lrns:
                for appr in alldataapprdist:
                    dd = []
            
                    for row in range(minrow,maxrow):
        
                        if row in invRows:
                            continue
                        if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):
                            continue
                        sr = str(row)
                        if findinlist(dicrows[sr]['dataappr'].lower().split('+'),appr)>=0 and findinlist(dicrows[sr]['lrnr'].lower().split(' '),lrnr)>=0 and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                           if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                                dd.append(dicrows[sr]['colVals'][ck])
                    if len(dd)>2:
                        ckdata[ck][lrnr+'--'+abbrdic[appr]] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Base Learner--Aggregated Data Approach Combination in Studies',vType,msList)



            #Compare levels of granularity
            ckdata[ck] = {}
            for dpl in alldpvarlevel:
                dd = []
            
                for row in range(minrow,maxrow):
        
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):               
                        continue
                    sr = str(row)
                    if dicrows[sr]['dpVarLevel'] == dpl and (dicrows[sr]['valueType'] == vType or dicrows[sr]['valueType']=='original'):
                       if ck in dicrows[sr]['colVals'] and dicrows[sr]['WC'] == 'C':
                            dd.append(dicrows[sr]['colVals'][ck])
                if len(dd)>2:
                    ckdata[ck][dpl] = dd
            if len(ckdata[ck].keys())>0:
                draw(ckdata,ck,'Level of Prediction in the Studies',vType,msList)


o = open('forestdata.txt','w')
o2= open('forestdata2.txt','w')
#Forest plots for studies

pprWandC={}

for msList in [[]]:#[],['precision','recall','f-measure'],['pf','recall','balance'],['f-measure'],['pf','recall','balance'],['precision','recall','f-measure'],['auc']]:
    
    for vType in sorted(list(allvalTypes)):

        for ms in msList if len(msList)>0 else colslst:
            
            cvsw = {}
        
            for row in range(minrow,maxrow):
                
                if row in invRows:
                    continue
                if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):
                    continue
                if dicrows[str(row)]['valueType'] != vType and dicrows[str(row)]['valueType'] != 'original':
                    continue
                if not ms in dicrows[str(row)]['colVals'].keys():
                    continue
                paper = dicrows[str(row)]['paper'] 
                if not paper in cvsw.keys():
                    cvsw[paper] = {}
                    cvsw[paper]['W'] = []
                    cvsw[paper]['C'] = []
            
                cvsw[paper][dicrows[str(row)]['WC']].append(dicrows[str(row)]['colVals'][ms])
            keys = list(cvsw.keys())
            for paper in keys:
                if len(cvsw[paper]['W'])<=1 or len(cvsw[paper]['C'])<=1:
                    del cvsw[paper]
            if len(cvsw.keys())==0:
                print(vType,ms,sorted(list(cvsw.keys())),'No paper')
                continue
            o2.write('#Raw Data for %s %s\n' % (vType,ms))
            Q = -1
            qpval = -1
            taw2 = -1

            
            if not ms in pprWandC.keys():
                pprWandC[ms] = {}
            if not vType in pprWandC[ms].keys():
                pprWandC[ms][vType] = []


            for paper in cvsw.keys():
                pprWandC[ms][vType].append(paper)
                meanW = np.mean(cvsw[paper]['W'])
                meanC = np.mean(cvsw[paper]['C'])
                stdW = np.std(cvsw[paper]['W'])
                stdC = np.std(cvsw[paper]['C'])
                nw = len(cvsw[paper]['W'])
                nc = len(cvsw[paper]['C'])
                
                stdP = math.sqrt(((nw-1)*stdW*stdW+(nc-1)*stdC*stdC)/(nw+nc-2))
                g = (meanC - meanW)/stdP
                d = J(nw+nc-2)*g
                nh = nhat(nw,nc)
                v = nh+((d*d)/(2*(nc+nw)))
                Ll = d-1.96*np.sqrt(v)
                Lu = d+1.96*np.sqrt(v)
                
                cvsw[paper]['meanW'] = meanW
                cvsw[paper]['meanC'] = meanC
                cvsw[paper]['stdW'] = stdW
                cvsw[paper]['stdC'] = stdC
                cvsw[paper]['stdP'] = stdP
                cvsw[paper]['nW'] = nw
                cvsw[paper]['nC'] = nc
                cvsw[paper]['n'] = nc+nw
                cvsw[paper]['g'] = g
                cvsw[paper]['d'] = d 
                cvsw[paper]['Ll'] = Ll
                cvsw[paper]['Lu'] = Lu
                cvsw[paper]['nh'] = nh
                cvsw[paper]['v'] = v
                cvsw[paper]['iw']=1.0/v
                cvsw[paper]['vsq'] = np.sqrt(v)
                o2.write('%s\t %f\t %f\t %d\t %f\t %f\t %d\n' % (paper,meanW,stdW,nw,meanC,stdC,nc))
            
            o2.write('\n\n\n\n')
            se = sum([1.0/cvsw[paper]['v'] for paper in cvsw.keys()])
            
            if se==0:
                print ('SE Zero - Fixed - Study',se, vType,ms)
                continue
            
            dstar = sum([cvsw[paper]['d']/cvsw[paper]['v'] for paper in cvsw.keys()])/se
            
            #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))                        
            dtaforest = []
            dtaforest2 = []
            dss = str(sorted(list(cvsw.keys())))
            
            
            frvals = []
            for i in range(50):
                if 'A'+str(i) in cvsw.keys():
                    paper = 'A'+str(i)
                    s = [paper,cvsw[paper]['d'],cvsw[paper]['Ll'],cvsw[paper]['Lu'],1.0/cvsw[paper]['v']]
                    frvals.append(cvsw[paper]['d'])
                    dtaforest.append(s)
                    o.write(str(s)+'\n')
            
            frsIndexes = sorted(range(len(frvals))) #, key=lambda k: frvals[k]
            
            
            
            s = ['@','InverseSum',vType,ms,dss]            
            dtaforest2.append(s)
            o.write(str(s)+'\n')
            s = ['Study','Center','Low','High','Weight']
            dtaforest2.append(s)
            o.write(str(s)+'\n')
            for fi,frsi in enumerate(frsIndexes):
                dtaforest2.append(dtaforest[frsi])
            dtaforest = dtaforest2

            s = ['Fixed',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
            dtaforest.append(s)
            o.write(str(s)+'\n')
            o.write('!\n!\n!\n!\n!\n')
            drawForstPlot(dtaforest,ms,'Forest Plot-InvSum',vType,msList,Q=Q,qpval=qpval,taw2=0)
            sfixed = s[:]
            print ('Fixed - Study',vType,ms)

            degf = len(cvsw.keys())-1
            Q = sum([((cvsw[paper]['d']-dstar)**2)/cvsw[paper]['v'] for paper in cvsw.keys()])
            qpval = chisqprob(Q,degf)
            taw2 = 0
            
            if Q<=degf:
                taw2 = 0                
            else:
                C = se - sum([(1.0/(cvsw[paper]['v'])**2) for paper in cvsw.keys()])/se
                if C==0:
                    print ('C Zero - Random - Study',C, vType,ms,se)
                    continue
                taw2 = (Q-degf)/C
           
            se = sum([1.0/(cvsw[paper]['v']+taw2) for paper in cvsw.keys()])
            if se==0:
                
                print ('SE Zero - Random - Study',se, vType,ms)
                continue

            dstar = sum([cvsw[paper]['d']/(cvsw[paper]['v']+taw2) for paper in cvsw.keys()])/se
            
            #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))
            fws = []
            dss = str(sorted(list(cvsw.keys())))
            for i in range(len(dtaforest)):
                if dtaforest[i][0] in cvsw.keys():
                    fws.append(dtaforest[i][-1])
                    dtaforest[i][-1] = 1.0/(cvsw[dtaforest[i][0]]['v']+taw2)
                    


            s = ['Random',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
            print ('Random - Study', vType,ms)
            dtaforest[-1]=s
            #o.write(str(s)+'\n')
            #o.write('!\n!\n!\n!\n!\n')
            drawForstPlot(dtaforest,ms,'Random-Forest Plot-InvSum',vType,msList,Q=Q,qpval=qpval,taw2=taw2,fws=fws,sfixed = sfixed)


#Forest plots for datasets and learners
for type in []:#[['ds','Dataset'],['lrnr','Learner']]:
    for msList in [[]]:#['f-measure'],['pf','recall','balance'],['precision','recall','f-measure'],['auc']]:
        for vType in sorted(list(allvalTypes)):
            for ms in msList if len(msList)>0 else colslst:
                cvsw = {}
                dss = set()
                for row in range(minrow,maxrow):
                
                    if row in invRows:
                        continue
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):
                        continue
                    if dicrows[str(row)]['valueType'] != vType and dicrows[str(row)]['valueType'] != 'original':
                        continue
                    if not ms in dicrows[str(row)]['colVals'].keys():
                        continue
                    ds = dicrows[str(row)][type[0]]
                    if ds.startswith('row') or ds.startswith('mn-') or ds.startswith('mdn-'):
                        continue
                    dss.add(ds)
                    if not ds in cvsw.keys():
                        cvsw[ds] = {}
                        cvsw[ds]['W'] = []
                        cvsw[ds]['C'] = []
                    cvsw[ds][dicrows[str(row)]['WC']].append(dicrows[str(row)]['colVals'][ms])
            
                
                keys = list(cvsw.keys())
                for ds in keys:
                    if len(cvsw[ds]['W'])<=1 or len(cvsw[ds]['C'])<=1:
                        del cvsw[ds]
                if len(cvsw.keys())==0:
                    print(vType,ms,sorted(list(cvsw.keys())),'No '+type[0])
                    continue

                o2.write('#Raw Data for %s %s %s\n' % (type[1],vType,ms))

                Q=-1
                qpval=-1
                taw2=-1
                for ds in cvsw.keys():
                    meanW = np.mean(cvsw[ds]['W'])
                    meanC = np.mean(cvsw[ds]['C'])
                    stdW = np.std(cvsw[ds]['W'])
                    stdC = np.std(cvsw[ds]['C'])
                    nw = len(cvsw[ds]['W'])
                    nc = len(cvsw[ds]['C'])
                    
                    stdP = math.sqrt(((nw-1)*stdW*stdW+(nc-1)*stdC*stdC)/(nw+nc-2))
                    
                    g = (meanC - meanW)/stdP
                    d = J(nw+nc-2)*g
                    nh = nhat(nw,nc)
                    v = nh+((d*d)/(2*(nc+nw)))
                    Ll = d-1.96*np.sqrt(v)
                    Lu = d+1.96*np.sqrt(v)
                    
                    cvsw[ds]['meanW'] = meanW
                    cvsw[ds]['meanC'] = meanC
                    cvsw[ds]['stdW'] = stdW
                    cvsw[ds]['stdC'] = stdC
                    cvsw[ds]['stdP'] = stdP
                    cvsw[ds]['nW'] = nw
                    cvsw[ds]['nC'] = nc
                    cvsw[ds]['n'] = nc+nw
                    cvsw[ds]['g'] = g
                    cvsw[ds]['d'] = d 
                    cvsw[ds]['Ll'] = Ll
                    cvsw[ds]['Lu'] = Lu
                    cvsw[ds]['nh'] = nh
                    cvsw[ds]['v'] = v
                    cvsw[ds]['vsq'] = np.sqrt(v)
                    
                    o2.write('%s\t %f\t %f\t %d\t %f\t %f\t %d\n' % (ds,meanW,stdW,nw,meanC,stdC,nc))
            
                o2.write('\n\n\n\n')
                se = sum([1.0/cvsw[ds]['v'] for ds in cvsw.keys()])
                if se==0:                    
                    print ('SE Zero - Fixed - ',type[0],se, vType,ms)
                    continue
                dstar = sum([cvsw[ds]['d']/cvsw[ds]['v'] for ds in cvsw.keys()])/se

                
                #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))
            
            
                dtaforest = []
                dtaforest2 = []
                dss = str(sorted(list(cvsw.keys())))
                
                frvals = []
                for ds in sorted(list(cvsw.keys())):
                    s = [ds,cvsw[ds]['d'],cvsw[ds]['Ll'],cvsw[ds]['Lu'],1.0/cvsw[ds]['v']]
                    dtaforest.append(s)
                    frvals.append(cvsw[ds]['d'])
                    o.write(str(s)+'\n')

                frsIndexes = sorted(range(len(frvals))) #, key=lambda k: frvals[k]

                s = ['@',type[1]+' InverseSum',vType,ms,dss]
                dtaforest2.append(s)
                o.write(str(s)+'\n')
                s = ['Study','Center','Low','High','Weight']
                dtaforest2.append(s)
                o.write(str(s)+'\n')
                for fi,frsi in enumerate(frsIndexes):
                    dtaforest2.append(dtaforest[frsi])
                dtaforest = dtaforest2


                if type[0]=='ds':
                    dtaforest2 = []
                    dtaforest2.append(dtaforest[0])
                    dtaforest2.append(dtaforest[1])
                    dtaforest[0] = None
                    dtaforest[1] = None
                    for so in suitesOrder:
                        for i in range(len(dtaforest)):
                            if dtaforest[i] == None:
                                continue
                            if dtaforest[i][0] in dsSuites[so]:
                                dtaforest2.append(dtaforest[i])
                                dtaforest[i] = None
                    for i in range(len(dtaforest)):
                        if dtaforest[i] == None:
                            continue
                        dtaforest2.append(dtaforest[i])
                    
                    if len(dtaforest)!=len(dtaforest2):
                        print ('Not Equal...Error')
                        input('')
                    dtaforest = dtaforest2
                
                s = ['Fixed',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
                dtaforest.append(s)
                o.write(str(s)+'\n')
                o.write('!\n!\n!\n!\n!\n')
                #sort based on dataset 
                print ('Fixed - ',type[0],vType,ms)                                                
                drawForstPlot(dtaforest,ms,'Forest Plot-'+type[1]+'-InvSum',vType,msList,Q=Q,qpval = qpval,taw2=taw2)
                sfixed = s[:]

                #Q = sum([((cvsw[ds]['d']-dstar)**2)/cvsw[ds]['v'] for ds in cvsw.keys()])
                #qpval = chisqprob(Q,len(cvsw.keys())-1)
                #print ('QPVAL->',Q,qpval)
                

                degf = len(cvsw.keys())-1
                Q = sum([((cvsw[ds]['d']-dstar)**2)/cvsw[ds]['v'] for ds in cvsw.keys()])
                qpval = chisqprob(Q,degf)
                taw2 = 0
            
                if Q<=degf:
                    taw2 = 0                
                else:
                    C = se - sum([(1.0/cvsw[ds]['v'])**2 for ds in cvsw.keys()])/se
                    if C==0:
                        print ('C Zero - Random - ',type[0],C, vType,ms)
                        continue
                    taw2 = (Q-degf)/C
           
                se = sum([1.0/(cvsw[ds]['v']+taw2) for ds in cvsw.keys()])
                if se==0:
                    
                    print ('SE Zero - Random - ',type[0],se, vType,ms)
                    continue
                dstar = sum([cvsw[ds]['d']/(cvsw[ds]['v']+taw2) for ds in cvsw.keys()])/se
            
                #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))
                fws = []
                dss = str(sorted(list(cvsw.keys())))
                for i in range(len(dtaforest)):                    
                    if dtaforest[i][0] in cvsw.keys():
                        fws.append(dtaforest[i][-1])
                        dtaforest[i][-1] = 1.0/(cvsw[dtaforest[i][0]]['v']+taw2)
                    


                s = ['Random',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
                print ('Random - ',type[0],vType,ms)
                dtaforest[-1]=s
                #o.write(str(s)+'\n')
                #o.write('!\n!\n!\n!\n!\n')
                drawForstPlot(dtaforest,ms,'Random-Forest Plot-'+type[1]+'-InvSum',vType,msList,Q=Q,qpval = qpval,taw2=taw2,fws=fws,sfixed=sfixed)

for type in [['ds','Dataset-SB'],['lrnr','Learner-SB']]:
    for msList in [[]]:#['f-measure'],['pf','recall','balance'],['precision','recall','f-measure'],['auc']]:
        for vType in sorted(list(allvalTypes)):
            for ms in msList if len(msList)>0 else colslst:
                cvsw = {}
                dss = set()
                for row in range(minrow,maxrow):
                    

                    if row in invRows:
                        continue
                    
                    if not set(dicrows[str(row)]['colVals'].keys()).issuperset(set(msList)):
                        continue
                    if dicrows[str(row)]['valueType'] != vType and dicrows[str(row)]['valueType'] != 'original':
                        continue
                    if not ms in dicrows[str(row)]['colVals'].keys():
                        continue
                    ds = dicrows[str(row)][type[0]]
                    if ds.startswith('row') or ds.startswith('mn-') or ds.startswith('mdn-'):
                        continue
                    paper = getpaper(row,papers)
                    if not ms in pprWandC.keys():
                        continue
                    if not vType in pprWandC[ms].keys():
                        print (ms,vType)
                        continue
                    if not paper in pprWandC[ms][vType]:
                        continue
                    dss.add(ds)
                    if not ds in cvsw.keys():
                        cvsw[ds] = {}
                        cvsw[ds]['W'] = []
                        cvsw[ds]['C'] = []
                    cvsw[ds][dicrows[str(row)]['WC']].append(dicrows[str(row)]['colVals'][ms])
            
                
                keys = list(cvsw.keys())
                for ds in keys:
                    if len(cvsw[ds]['W'])<=1 or len(cvsw[ds]['C'])<=1:
                        del cvsw[ds]
                if len(cvsw.keys())==0:
                    print(vType,ms,sorted(list(cvsw.keys())),'No '+type[0])
                    continue

                o2.write('#Raw Data for %s %s %s\n' % (type[1],vType,ms))

                Q=-1
                qpval=-1
                taw2=-1
                for ds in cvsw.keys():
                    meanW = np.mean(cvsw[ds]['W'])
                    meanC = np.mean(cvsw[ds]['C'])
                    stdW = np.std(cvsw[ds]['W'])
                    stdC = np.std(cvsw[ds]['C'])
                    nw = len(cvsw[ds]['W'])
                    nc = len(cvsw[ds]['C'])
                    
                    stdP = math.sqrt(((nw-1)*stdW*stdW+(nc-1)*stdC*stdC)/(nw+nc-2))
                    
                    g = (meanC - meanW)/stdP
                    d = J(nw+nc-2)*g
                    nh = nhat(nw,nc)
                    v = nh+((d*d)/(2*(nc+nw)))
                    Ll = d-1.96*np.sqrt(v)
                    Lu = d+1.96*np.sqrt(v)
                    
                    cvsw[ds]['meanW'] = meanW
                    cvsw[ds]['meanC'] = meanC
                    cvsw[ds]['stdW'] = stdW
                    cvsw[ds]['stdC'] = stdC
                    cvsw[ds]['stdP'] = stdP
                    cvsw[ds]['nW'] = nw
                    cvsw[ds]['nC'] = nc
                    cvsw[ds]['n'] = nc+nw
                    cvsw[ds]['g'] = g
                    cvsw[ds]['d'] = d 
                    cvsw[ds]['Ll'] = Ll
                    cvsw[ds]['Lu'] = Lu
                    cvsw[ds]['nh'] = nh
                    cvsw[ds]['v'] = v
                    cvsw[ds]['vsq'] = np.sqrt(v)
                    
                    o2.write('%s\t %f\t %f\t %d\t %f\t %f\t %d\n' % (ds,meanW,stdW,nw,meanC,stdC,nc))
            
                o2.write('\n\n\n\n')
                se = sum([1.0/cvsw[ds]['v'] for ds in cvsw.keys()])
                if se==0:                    
                    print ('SE Zero - Fixed - ',type[0],se, vType,ms)
                    continue
                dstar = sum([cvsw[ds]['d']/cvsw[ds]['v'] for ds in cvsw.keys()])/se

                
                #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))
            
            
                dtaforest = []
                dtaforest2 = []
                dss = str(sorted(list(cvsw.keys())))
                
                frvals = []
                for ds in sorted(list(cvsw.keys())):
                    s = [ds,cvsw[ds]['d'],cvsw[ds]['Ll'],cvsw[ds]['Lu'],1.0/cvsw[ds]['v']]
                    dtaforest.append(s)
                    frvals.append(cvsw[ds]['d'])
                    o.write(str(s)+'\n')

                frsIndexes = sorted(range(len(frvals))) #, key=lambda k: frvals[k]

                s = ['@',type[1]+' InverseSum',vType,ms,dss]
                dtaforest2.append(s)
                o.write(str(s)+'\n')
                s = ['Study','Center','Low','High','Weight']
                dtaforest2.append(s)
                o.write(str(s)+'\n')
                for fi,frsi in enumerate(frsIndexes):
                    dtaforest2.append(dtaforest[frsi])
                dtaforest = dtaforest2


                if type[0]=='ds':
                    dtaforest2 = []
                    dtaforest2.append(dtaforest[0])
                    dtaforest2.append(dtaforest[1])
                    dtaforest[0] = None
                    dtaforest[1] = None
                    for so in suitesOrder:
                        for i in range(len(dtaforest)):
                            if dtaforest[i] == None:
                                continue
                            if dtaforest[i][0] in dsSuites[so]:
                                dtaforest2.append(dtaforest[i])
                                dtaforest[i] = None
                    for i in range(len(dtaforest)):
                        if dtaforest[i] == None:
                            continue
                        dtaforest2.append(dtaforest[i])
                    
                    if len(dtaforest)!=len(dtaforest2):
                        print ('Not Equal...Error')
                        input('')
                    dtaforest = dtaforest2
                
                s = ['Fixed',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
                dtaforest.append(s)
                o.write(str(s)+'\n')
                o.write('!\n!\n!\n!\n!\n')
                #sort based on dataset 
                print ('Fixed - ',type[0],vType,ms)                                                
                drawForstPlot(dtaforest,ms,'Forest Plot-'+type[1]+'-InvSum',vType,msList,Q=Q,qpval = qpval,taw2=taw2)
                sfixed = s[:]

                #Q = sum([((cvsw[ds]['d']-dstar)**2)/cvsw[ds]['v'] for ds in cvsw.keys()])
                #qpval = chisqprob(Q,len(cvsw.keys())-1)
                #print ('QPVAL->',Q,qpval)
                

                degf = len(cvsw.keys())-1
                Q = sum([((cvsw[ds]['d']-dstar)**2)/cvsw[ds]['v'] for ds in cvsw.keys()])
                qpval = chisqprob(Q,degf)
                taw2 = 0
            
                if Q<=degf:
                    taw2 = 0                
                else:
                    C = se - sum([(1.0/cvsw[ds]['v'])**2 for ds in cvsw.keys()])/se
                    if C==0:
                        print ('C Zero - Random - ',type[0],C, vType,ms)
                        continue
                    taw2 = (Q-degf)/C
           
                se = sum([1.0/(cvsw[ds]['v']+taw2) for ds in cvsw.keys()])
                if se==0:
                    
                    print ('SE Zero - Random - ',type[0],se, vType,ms)
                    continue
                dstar = sum([cvsw[ds]['d']/(cvsw[ds]['v']+taw2) for ds in cvsw.keys()])/se
            
                #print(vType,ms,sorted(list(cvsw.keys())),dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se))
                fws = []
                dss = str(sorted(list(cvsw.keys())))
                for i in range(len(dtaforest)):                    
                    if dtaforest[i][0] in cvsw.keys():
                        fws.append(dtaforest[i][-1])
                        dtaforest[i][-1] = 1.0/(cvsw[dtaforest[i][0]]['v']+taw2)
                    


                s = ['Random',dstar,dstar-1.96*math.sqrt(1/se),dstar+1.96*math.sqrt(1/se),se,math.sqrt(1/se),dstar/math.sqrt(1/se),2*norm.sf(abs(dstar/math.sqrt(1/se)))]
                print ('Random - ',type[0],vType,ms)
                dtaforest[-1]=s
                #o.write(str(s)+'\n')
                #o.write('!\n!\n!\n!\n!\n')
                drawForstPlot(dtaforest,ms,'Random-Forest Plot-'+type[1]+'-InvSum',vType,msList,Q=Q,qpval = qpval,taw2=taw2,fws=fws,sfixed=sfixed)


#List of the papers for each reporting type
pprsvTypes = {}
for vType in allvalTypes:
    pprsvTypes[vType] = set()
    for row in range(minrow,maxrow):
        if row in invRows:
            continue
        if dicrows[str(row)]['valueType'].lower() == vType.lower() and dicrows[str(row)]['WC'] == 'C':
            pprsvTypes[vType].add(dicrows[str(row)]['paper'])
    print (vType, sorted(list(pprsvTypes[vType])))


#List of the papers that use multiple ways of reporting
multivType = set()
for p in papers.keys():
    count = 0
    for vType in allvalTypes:
        if p in pprsvTypes[vType]:
            count+=1
    if count>1:
        multivType.add(p)
print ('Multi', sorted(list(multivType)))




#List of papers for each metric suite or combination
print ('\n\n\n Metric Types')
metricTypes = {}
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    if dicrows[str(row)]['WC'] == 'C':
        mtr = dicrows[str(row)]['metrics']
        if not mtr in metricTypes.keys():
            metricTypes[mtr] = [1,set([dicrows[str(row)]['paper']])]
        else:
            metricTypes[mtr][0]+=1
            metricTypes[mtr][1].add(dicrows[str(row)]['paper'])

for mtr in metricTypes.keys():    
    print (mtr,metricTypes[mtr][0],sorted(list(metricTypes[mtr][1])))   

multimtype = set()
for p in papers.keys():

    
    count = 0
    for mtr in metricTypes.keys():    
        if p in metricTypes[mtr][1]:
            count+=1
    if count>1:
        multimtype.add(p)
print (sorted(list(multimtype)))

print('Papers in only one cat:')
for mtr in metricTypes.keys():
    if mtr.find('+')>0:
        continue
    for p in metricTypes[mtr][1]:
        if not p in multimtype:
            print (p)

print('\n\n')
for s in ['oo','loc','scm','process']:
    tset = set()
    for mtr in metricTypes.keys():
        if findinlist(mtr.lower().split('+'),s)>=0:        
            for p in metricTypes[mtr][1]:
                tset.add(p)
    print(s,sorted(list(tset)),len(tset))

print ('\n\n\n')

#Learners and their usage statistics in the studies
lrnrscounts = {}
lrnrsStudies = {}
for row in range(minrow,maxrow):
    if row in invRows:
        continue
    if dicrows[str(row)]['WC'] == 'C':
        lrnr = dicrows[str(row)]['lrnr']
        if lrnr in lrnrsStudies:
            lrnrsStudies[lrnr].add(dicrows[str(row)]['paper'])
        else:
            lrnrsStudies[lrnr] = set([dicrows[str(row)]['paper']])
        if lrnr in lrnrscounts.keys():
            lrnrscounts[lrnr]+=1
        else:
            lrnrscounts[lrnr]=1
for lrnr in lrnrscounts.keys():
    print (lrnr,lrnrscounts[lrnr])

print ('\n\n\n')
import operator
lrnrsc = sorted(lrnrscounts.items(),key = lambda x: x[1])
print (lrnrsc,'\n\n\n\n')

baselrn = {}

for lrn in lrnrsStudies.keys():
    lrn = lrn.strip()

    if lrn.find(' ')>0:
        lrn = lrn[lrn.find(' '):]
    baselrn[lrn] = set()

for lrn in lrnrsStudies.keys():
    lrn = lrn.strip()
    lrn2 = lrn
    if lrn.find(' ')>0:
        lrn2 = lrn[lrn.find(' '):]
    for p in lrnrsStudies[lrn]:
        baselrn[lrn2].add(p)

for l in lrnrsc:
    lrnr = l[0]
    if lrnr in baselrn.keys():    
        print (lrnr,l[1],baselrn[lrnr],len(baselrn[lrnr]))

print('\n\n\n\n\n')



#Data approach and studies
for appr in alldataapprdist:
    tset = set()
    for row in range(minrow,maxrow):
        if row in invRows:
            continue

        rowappr = dicrows[str(row)]['dataappr']
        if findinlist(rowappr.lower().split('+'),appr)>=0:
            p = dicrows[str(row)]['paper']
            tset.add(p)
    print(appr,sorted(list(tset)))

print ('\n\n\n')





#Select the plots for the study
from shutil import copyfile

def copyFilesForPaper():
    o = open('vpl.txt')
    lines = o.readlines()
    o.close()
    if not os.path.exists(basePath+'CopyData/V/'):
        os.makedirs(basePath+'CopyData/V/')
    if not os.path.exists(basePath+'jpg/CopyData/V/'):
        os.makedirs(basePath+'jpg/CopyData/V/')
    

    for line in lines:
        if line.endswith('\n'):
            line = line[:-1]
        if len(line)<5:
            continue
        
        if os.path.exists(basePath+'allin1/'+line):
            copyfile(basePath+'allin1/'+line, basePath+'CopyData/V/'+line)
        else:
            print ('File Not Found:' , line)

        if os.path.exists(basePath+'jpg/allin1/'+line[:-4]+'.jpg'):
            copyfile(basePath+'jpg/allin1/'+line[:-4]+'.jpg', basePath+'jpg/CopyData/V/'+line[:-4]+'.jpg')
        else:
            print ('File Not Found:' , line[:-4]+'.jpg')


    o = open('fpl.txt')
    lines = o.readlines()
    o.close()
    if not os.path.exists(basePath+'CopyData/F/'):
        os.makedirs(basePath+'CopyData/F/')
    if not os.path.exists(basePath+'jpg/CopyData/F/'):
        os.makedirs(basePath+'jpg/CopyData/F/')


    for line in lines:
        if line.endswith('\n'):
            line = line[:-1]
        if len(line)<5:
            continue
        
        if os.path.exists(basePath+'allin1Forest/'+line):
            copyfile(basePath+'allin1Forest/'+line, basePath+'CopyData/F/'+line)
        else:
            print ('File Not Found:' , line)

        if os.path.exists(basePath+'jpg/allin1Forest/'+line[:-4]+'.jpg'):
            copyfile(basePath+'jpg/allin1Forest/'+line[:-4]+'.jpg', basePath+'jpg/CopyData/F/'+line[:-4]+'.jpg')
        else:
            print ('File Not Found:' , line[:-4]+'.jpg')


copyFilesForPaper()

for c in cols.keys():
    print (c)
    print (len(evcriteria[c][0]),sorted (list(evcriteria[c][0])))
    print (len(evcriteria[c][1]),sorted (list(evcriteria[c][1])))
 
o = open('evplotdata.txt','w')

for c in cols.keys():
    o.write ('%s\t%d\t%d\n' % (c,len(evcriteria[c][0]),len(evcriteria[c][1])))
o.close()